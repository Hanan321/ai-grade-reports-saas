from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from config.default_config import ReportBrandingConfig
from engine.exports import export_excel_workbook


def test_export_excel_workbook_includes_product_sheets_and_formatting():
    cleaned = pd.DataFrame(
        {
            "student_id": [1, 2, 3],
            "student_name": ["Ali", "Aisha", "Sara"],
            "subject": ["Math", "Math", "Science"],
            "test_date": pd.to_datetime(["2025-01-01", "2025-01-02", "2025-01-03"]),
            "homework": [60, 90, 80],
            "quizscore": [65, 92, 84],
            "exam_score": [66, 95, 88],
            "attendance_percent": [75, 98, 70],
            "final_score": [64.0, 92.9, 84.8],
            "low_attendance": [True, False, True],
            "at_risk": [True, False, True],
        }
    )
    student_summary = pd.DataFrame(
        {
            "student_name": ["Aisha", "Sara", "Ali"],
            "tests_taken": [1, 1, 1],
            "avg_final_score": [92.9, 84.8, 64.0],
            "avg_attendance": [98.0, 70.0, 75.0],
            "any_risk": [False, True, True],
        }
    )
    subject_summary = pd.DataFrame(
        {
            "subject": ["Math", "Science"],
            "tests_taken": [2, 1],
            "avg_final_score": [78.5, 84.8],
            "avg_attendance": [86.5, 70.0],
        }
    )

    workbook_bytes = export_excel_workbook(
        cleaned,
        student_summary,
        subject_summary,
        warnings=["1 row has a missing score."],
    )
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == [
        "Cleaned Data",
        "Summary By Student",
        "Summary By Subject",
        "At-Risk Students",
        "High-Performing Students",
        "Low Attendance Students",
        "Validation Summary",
    ]
    assert workbook["Cleaned Data"].freeze_panes == "A2"
    assert workbook["Cleaned Data"]["A1"].font.bold is True
    assert workbook["At-Risk Students"].max_row == 3
    assert workbook["High-Performing Students"].max_row == 2
    assert workbook["Low Attendance Students"].max_row == 3


def test_export_excel_workbook_applies_report_branding_metadata():
    cleaned = pd.DataFrame({"student_name": ["Ali"], "final_score": [90], "at_risk": [False]})
    student_summary = pd.DataFrame({"student_name": ["Ali"], "avg_final_score": [90]})
    subject_summary = pd.DataFrame({"subject": ["Math"], "avg_final_score": [90]})
    branding = ReportBrandingConfig(
        report_title="Demo School Workbook",
        header_text="Official report",
        footer_text="Confidential",
    )

    workbook_bytes = export_excel_workbook(
        cleaned,
        student_summary,
        subject_summary,
        report_branding=branding,
    )
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.properties.title == "Demo School Workbook"
    assert workbook.properties.subject == "Official report"
    assert workbook.properties.description == "Confidential"
