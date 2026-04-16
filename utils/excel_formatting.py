"""Professional workbook formatting helpers."""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill("solid", fgColor="F8FAFC")
HEADER_FONT = Font(color="111827", bold=True)
AT_RISK_FILL = PatternFill("solid", fgColor="FFF1F2")
HIGH_PERFORMER_FILL = PatternFill("solid", fgColor="DCFCE7")
LOW_ATTENDANCE_FILL = PatternFill("solid", fgColor="FEF3C7")


def format_report_workbook(workbook, high_performer_threshold: float = 85.0) -> None:
    """Apply shared formatting to every worksheet in the report workbook."""

    for worksheet in workbook.worksheets:
        format_report_sheet(worksheet, high_performer_threshold=high_performer_threshold)


def format_report_sheet(worksheet: Worksheet, high_performer_threshold: float = 85.0) -> None:
    """Style a worksheet header, sizing, panes, and relevant status rows."""

    if worksheet.max_row == 0:
        return

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    _style_header_row(worksheet)
    _adjust_column_widths(worksheet)
    _format_dates_and_numbers(worksheet)
    _apply_status_row_fills(worksheet, high_performer_threshold=high_performer_threshold)
    _apply_score_conditional_formatting(worksheet, high_performer_threshold=high_performer_threshold)


def _style_header_row(worksheet: Worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 22


def _adjust_column_widths(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 36)


def _format_dates_and_numbers(worksheet: Worksheet) -> None:
    headers = _header_lookup(worksheet)
    for header, column_index in headers.items():
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for value_cell in cell:
                if header == "test_date":
                    value_cell.number_format = "yyyy-mm-dd"
                elif "score" in header or "attendance" in header:
                    value_cell.number_format = "0.0"


def _apply_status_row_fills(worksheet: Worksheet, high_performer_threshold: float) -> None:
    headers = _header_lookup(worksheet)
    at_risk_column = headers.get("at_risk") or headers.get("any_risk")
    high_score_column = headers.get("final_score") or headers.get("avg_final_score")
    low_attendance_column = headers.get("low_attendance")

    for row_index in range(2, worksheet.max_row + 1):
        fill = None
        if at_risk_column and worksheet.cell(row_index, at_risk_column).value is True:
            fill = AT_RISK_FILL
        elif low_attendance_column and worksheet.cell(row_index, low_attendance_column).value is True:
            fill = LOW_ATTENDANCE_FILL
        elif high_score_column:
            score = worksheet.cell(row_index, high_score_column).value
            if isinstance(score, (int, float)) and score >= high_performer_threshold:
                fill = HIGH_PERFORMER_FILL

        if fill:
            for cell in worksheet[row_index]:
                cell.fill = fill


def _apply_score_conditional_formatting(worksheet: Worksheet, high_performer_threshold: float) -> None:
    headers = _header_lookup(worksheet)
    for header in ("final_score", "avg_final_score"):
        column_index = headers.get(header)
        if not column_index or worksheet.max_row < 2:
            continue
        letter = get_column_letter(column_index)
        cell_range = f"{letter}2:{letter}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["70"], fill=AT_RISK_FILL),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(high_performer_threshold)],
                fill=HIGH_PERFORMER_FILL,
            ),
        )


def _header_lookup(worksheet: Worksheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
